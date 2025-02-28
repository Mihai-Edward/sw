import pandas as pd
from datetime import datetime
import os
import numpy as np
from openpyxl import Workbook, load_workbook
from lottery_ml_model import LotteryMLPredictor  # Updated import

def save_draw_to_csv(draw_date, draw_numbers, csv_file='C:\\Users\\MihaiNita\\OneDrive - Prime Batteries\\Desktop\\proiectnow\\Versiune1.4\\src\\historical_draws.csv'):
    """
    Save draw numbers to CSV file and display them in console
    
    Args:
        draw_date: Date/time of the draw
        draw_numbers: List of numbers from the draw
        csv_file: Path to the CSV file
    """
    # Display in console
    print(f"\nSaving draw to ML database: {draw_date}")
    print("Numbers:", ', '.join(map(str, sorted(draw_numbers))))
    
    # Ensure draw_numbers has exactly 20 numbers
    if len(draw_numbers) > 20:
        draw_numbers = sorted(draw_numbers)[:20]
    elif len(draw_numbers) < 20:
        print("Warning: Draw has less than 20 numbers")
        return
    
    # Prepare data for CSV
    draw_dict = {f'number{i+1}': num for i, num in enumerate(sorted(draw_numbers))}
    draw_dict['date'] = draw_date
    
    # Check if file exists
    if not os.path.exists(csv_file):
        # Create new DataFrame with headers
        df = pd.DataFrame([draw_dict])
    else:
        # Read existing file
        try:
            df = pd.read_csv(csv_file)
            # Check if this draw already exists
            if not df.empty and 'date' in df.columns:
                if draw_date in df['date'].values:
                    print(f"Draw for {draw_date} already exists in database")
                    return
            df = pd.concat([df, pd.DataFrame([draw_dict])], ignore_index=True)
        except Exception as e:
            print(f"Error reading existing CSV: {str(e)}")
            df = pd.DataFrame([draw_dict])
    
    # Save to CSV
    try:
        df.to_csv(csv_file, index=False)
        print(f"Successfully saved draw to {csv_file}")
    except Exception as e:
        print(f"Error saving to CSV: {str(e)}")

def save_predictions_to_excel(predictions, probabilities, timestamp, excel_file='data/processed/predictions.xlsx'):
    """
    Save ML predictions and their probabilities to an Excel file in the 'data/processed' directory
    
    Args:
        predictions: List of predicted numbers
        probabilities: List of probabilities corresponding to the predicted numbers
        timestamp: Time when the prediction was made
        excel_file: Path to the Excel file
    """
    # Prepare data for Excel
    data = {
        'Timestamp': [timestamp] * len(predictions),
        'Number': predictions,
        'Probability': probabilities
    }
    df = pd.DataFrame(data)

    if os.path.exists(excel_file):
        # Load existing workbook and sheet
        book = load_workbook(excel_file)
        writer = pd.ExcelWriter(excel_file, engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        
        # Get the last row in the existing Excel sheet
        startrow = writer.sheets['Predictions'].max_row
        
        # Append the new data
        df.to_excel(writer, sheet_name='Predictions', index=False, header=False, startrow=startrow)
        
        # Save the workbook
        writer.close()
    else:
        # Create a new workbook and sheet
        book = Workbook()
        sheet = book.active
        sheet.title = 'Predictions'
        
        # Write the data
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Predictions', index=False)

    print(f"\nPredictions saved to {excel_file}")

def train_ml_models(csv_file='C:\\Users\\MihaiNita\\OneDrive - Prime Batteries\\Desktop\\proiectnow\\Versiune1.4\\src\\historical_draws.csv', models_dir='src/ml_models'):
    """
    Train ML models based on historical data
    
    Args:
        csv_file: Path to the CSV file containing historical data
        models_dir: Directory to save the trained models
    """
    try:
        # Check if CSV file exists
        if not os.path.exists(csv_file):
            print(f"\nError: No historical data found at {csv_file}")
            return None
        
        # Load historical data
        historical_data = pd.read_csv(csv_file)
        
        # Verify data format
        required_columns = ['date'] + [f'number{i}' for i in range(1, 21)]
        missing_columns = [col for col in required_columns if col not in historical_data.columns]
        if missing_columns:
            print(f"\nError: Missing columns in data: {missing_columns}")
            return None
        
        # Ensure datetime conversion and feature extraction
        historical_data['date'] = pd.to_datetime(historical_data['date'], errors='coerce')
        historical_data['day_of_week'] = historical_data['date'].dt.dayofweek
        historical_data['month'] = historical_data['date'].dt.month
        historical_data['day_of_year'] = historical_data['date'].dt.dayofyear
        historical_data['days_since_first_draw'] = (historical_data['date'] - historical_data['date'].min()).dt.days
        
        # Check if we have enough historical data (need at least 6 draws)
        if len(historical_data) < 6:
            print(f"\nNot enough historical data. Need at least 6 draws, but only have {len(historical_data)}.")
            return None
        
        print(f"\nFound {len(historical_data)} historical draws in database.")
        
        # Initialize predictor
        predictor = LotteryMLPredictor(numbers_range=(1, 80), numbers_to_draw=20)
        
        print("Training new ML models...")
        # Prepare and train models
        X, y = predictor.prepare_data(historical_data)
        predictor.train_models(X, y)
        
        # Save models
        os.makedirs(models_dir, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        predictor.save_models(f'{models_dir}/lottery_predictor_{timestamp}')
        print("Models trained and saved successfully!")
        
    except Exception as e:
        print(f"\nError in training ML models: {str(e)}")

def get_ml_prediction(csv_file='C:\\Users\\MihaiNita\\OneDrive - Prime Batteries\\Desktop\\proiectnow\\Versiune1.4\\src\\historical_draws.csv'):
    """
    Get prediction from ML model based on historical data
    """
    try:
        # Check if CSV file exists
        if not os.path.exists(csv_file):
            print(f"\nError: No historical data found at {csv_file}")
            print("Please use option 3 first to collect draws.")
            return None
            
        # Load historical data
        historical_data = pd.read_csv(csv_file)
        
        # Ensure datetime conversion and feature extraction
        historical_data['date'] = pd.to_datetime(historical_data['date'], errors='coerce')
        historical_data['day_of_week'] = historical_data['date'].dt.dayofweek
        historical_data['month'] = historical_data['date'].dt.month
        historical_data['day_of_year'] = historical_data['date'].dt.dayofyear
        historical_data['days_since_first_draw'] = (historical_data['date'] - historical_data['date'].min()).dt.days
        
        # Initialize predictor
        predictor = LotteryMLPredictor(numbers_range=(1, 80), numbers_to_draw=20)
        
        # Load existing models if available
        models_dir = 'src/ml_models'
        model_base = None
        
        # Try to read the timestamp from the file first
        if os.path.exists('src/model_timestamp.txt'):
            try:
                with open('src/model_timestamp.txt', 'r') as f:
                    timestamp = f.read().strip()
                    model_base = f'{models_dir}/lottery_predictor_{timestamp}'
                    print(f"Loading model from {model_base}")
            except Exception as e:
                print(f"Error reading timestamp file: {str(e)}")
        
        # If no timestamp file or model not found, try to find the latest model file
        if not model_base:
            print("No timestamp file found, searching for latest model...")
            model_files = [f for f in os.listdir(models_dir) if f.endswith('_model.pkl')]
            if model_files:
                latest = max(model_files)
                model_base = os.path.join(models_dir, latest.replace('_model.pkl', ''))
                print(f"Found latest model: {model_base}")
        
        if model_base and os.path.exists(f"{model_base}_pattern_model.pkl"):
            print("Loading existing ML models...")
            predictor.load_models(model_base)
            
            # Prepare recent draws for prediction
            number_cols = [f'number{i}' for i in range(1, 21)]
            recent_draws = historical_data.tail(5)[number_cols + ['date', 'day_of_week', 'month', 'day_of_year', 'days_since_first_draw']]
            
            # Ensure recent_draws has the correct shape
            if recent_draws.shape[1] != 25:  # 20 number columns + 5 date feature columns
                raise ValueError("Recent draws data shape is incorrect. Expected 25 columns.")
            
            # Generate prediction
            predicted_numbers, probabilities = predictor.predict(recent_draws)
            
            # Display prediction results
            print("\n=== ML Prediction Results ===")
            print("Predicted numbers for next draw:", sorted(predicted_numbers))
            
            print("\nTop 10 most likely numbers and their probabilities:")
            number_probs = [(number, prob) for number, prob in zip(range(1, 81), probabilities)]
            number_probs.sort(key=lambda x: x[1], reverse=True)
            
            # Save top 10 predictions to Excel
            predictions_dir = 'src/data/processed'
            os.makedirs(predictions_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            predictions_data = {
                'Timestamp': [timestamp] * 10,
                'Number': [num for num, _ in number_probs[:10]],
                'Probability': [prob for _, prob in number_probs[:10]]
            }
            
            predictions_df = pd.DataFrame(predictions_data)
            predictions_file = os.path.join(predictions_dir, 'predictions.xlsx')
            
            if os.path.exists(predictions_file):
                with pd.ExcelWriter(predictions_file, mode='a', engine='openpyxl', 
                                    if_sheet_exists='overlay') as writer:
                    predictions_df.to_excel(writer, sheet_name='Predictions', 
                                            header=not os.path.exists(predictions_file),
                                            index=False)
            else:
                predictions_df.to_excel(predictions_file, sheet_name='Predictions', index=False)
            
            print("\nPrediction results:")
            for number, prob in number_probs[:10]:
                print(f"Number {number:2d}: {prob:.4f}")
            
            print(f"\nPredictions saved to {predictions_file}")
            
            return predicted_numbers
        else:
            print(f"\nError: No valid model files found in {models_dir}")
            return None
            
    except Exception as e:
        print(f"\nError in ML prediction: {str(e)}")
        print("\nDebug information:")
        print(f"CSV file exists: {os.path.exists(csv_file)}")
        if os.path.exists(csv_file):
            try:
                df = pd.read_csv(csv_file)
                print(f"Number of rows in data: {len(df)}")
                print("Columns in data:", df.columns.tolist())
                print("\nFirst few rows of data:")
                print(df.head())
            except Exception as read_error:
                print(f"Error reading CSV file: {str(read_error)}")
        return None